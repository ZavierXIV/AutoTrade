#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import os
from openpyxl import Workbook


# In[2]:


mydir = 'C:/Users/Alex/desktop/Python/e/'


# In[3]:


wb = Workbook()


# In[4]:


wb.create_sheet('sheet1', index=1)
wb.create_sheet('sheet2', index=2)
wb.create_sheet('sheet3', index=3)


# In[5]:


sheet1 = wb['sheet1']
sheet2 = wb['sheet2']
sheet3 = wb['sheet3']


# In[6]:


FileList = os.listdir(mydir)


# In[7]:


def setTitle(datas,sheet):
    if((datas[0][0][3] in array) == False):
        array.append(datas[0][0][3])
        datas = (datas[0][0]).drop([0,1,2], axis=0)
        tmplist = datas.tolist()
        sheet.append(tmplist)


# In[8]:


array = []
for i in FileList:
    mydir1 = mydir + i + '/'
    FileList1 = os.listdir(mydir1)
    for n in FileList1:
        if( n.find('html') != -1):
            htmls = pd.read_html(mydir1 + n)
            html = (htmls[0][1]).drop([0,1,2], axis=0)
            list = html.tolist()
            if '標示部' in html[3] != -1:
                tarray = setTitle(htmls,sheet1)
                sheet1.append(list)
            elif '所有權部' in html[3]:
                tarray = setTitle(htmls,sheet2)
                sheet2.append(list)
            elif '權利部' in html[3]:
                tarray = setTitle(htmls,sheet3)
                sheet3.append(list)
            else:
                print(n)
                break;


# In[9]:


wb.save(r'Data.xlsx')

